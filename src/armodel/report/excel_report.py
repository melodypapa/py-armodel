import logging

from openpyxl import Workbook


class ExcelReporter:
    def __init__(self) -> None:
        self.wb = Workbook()
        self._logger = logging.getLogger()

    def write_revision(self):
        sheet = self.wb['Sheet']
        sheet.title = "History"

        title_rows = ["When", "Who", "Version", "History"]
        self.write_title_row(sheet, title_rows)

    def auto_width(self, worksheet):
        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))

        for col, value in dims.items():
            worksheet.column_dimensions[col].width = (value + 2) + 2

    def write_title_row(self, sheet, title_row):
        for idx in range(0, len(title_row)):
            cell = sheet.cell(row=1, column=idx + 1)
            cell.value = title_row[idx]

    def write_cell(self, sheet, row, column, value, format = None):
        cell = sheet.cell(row = row, column=column)
        cell.value = value
        if (format != None):
            if ('alignment' in format):
                cell.alignment = format['alignment']
            if ('number_format' in format):
                cell.number_format = format['number_format']

    def save(self, name: str):
        self.wb.save(name)
