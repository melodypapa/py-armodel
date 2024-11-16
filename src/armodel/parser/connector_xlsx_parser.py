from typing import Dict, List
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

import openpyxl
import re

from ..models.M2.autosar_templates.generic_structure.ar_package import ARPackage

from ..parser.excel_parser import AbstractExcelParser
from ..data_models.sw_connector import AssemblySwConnectorData, DelegationSwConnectorData, SwConnectorData
from ..models.ar_ref import PPortInCompositionInstanceRef, RPortInCompositionInstanceRef, RefType
from ..models.sw_component import CompositionSwComponentType
from ..models.M2.autosar_templates.autosar_top_level_structure import AUTOSAR

class ConnectorXls:

    COL_SHORT_NAME      = 'Short Name'
    COL_INNER_SW_C      = 'Inner SW-C'
    COL_INNER_PPORT     = 'Inner PPort'
    COL_OUTER_PPORT     = 'Outer PPort'
    COL_INNER_RPORT     = 'Inner RPort'
    COL_OUTER_RPORT     = 'Outer RPort'

    COL_PROVIDER_SW_C    = "Provide SW-C"
    COL_PPORT           = "PPort"
    COL_REQUESTER_SW_C    = "Request SW-C"
    COL_RPORT           = "RPort"

class ConnectorXlsReader(AbstractExcelParser):
    def __init__(self) -> None:
        super().__init__()

        self.column_delegation_sw_connectors = {
            ConnectorXls.COL_SHORT_NAME: -1,
            ConnectorXls.COL_INNER_SW_C: -1,
            ConnectorXls.COL_INNER_PPORT: -1,
            ConnectorXls.COL_OUTER_PPORT: -1,
            ConnectorXls.COL_INNER_RPORT: -1,
            ConnectorXls.COL_OUTER_RPORT: -1,
        }

        self.column_assembly_sw_connectors = {
            ConnectorXls.COL_SHORT_NAME: -1,
            ConnectorXls.COL_PROVIDER_SW_C: -1,
            ConnectorXls.COL_PPORT: -1,
            ConnectorXls.COL_REQUESTER_SW_C: -1,
            ConnectorXls.COL_RPORT: -1,
        }

        self.sw_connectors = {}    # type: Dict[str, List[DelegationSwConnectorData]]
        

    def getCompositionSwComponentList(self) -> List[str]:
        return self.sw_connectors.keys()
            
    def getSwConnectorList(self, swc: str) -> List[SwConnectorData]:
        if swc not in self.sw_connectors:
            self.sw_connectors[swc] = []
        return self.sw_connectors[swc]
        #return sorted(self.sw_connectors[swc], key = lambda o: o.short_name)
            
    def readDelegationSwConnectors(self, sheet: Worksheet, swc: str, start_row: int, column_list: Dict[str, int]):
        connectors = self.getSwConnectorList(swc)
        for row in sheet.iter_rows(min_row = start_row, values_only= True):
            connector = DelegationSwConnectorData()
            connector.short_name = row[column_list[ConnectorXls.COL_SHORT_NAME]]
            connector.inner_swc = row[column_list[ConnectorXls.COL_INNER_SW_C]]
            connector.inner_pport = row[column_list[ConnectorXls.COL_INNER_PPORT]]
            connector.inner_rport = row[column_list[ConnectorXls.COL_INNER_RPORT]]
            connector.outer_pport = row[column_list[ConnectorXls.COL_OUTER_PPORT]]
            connector.outer_rport = row[column_list[ConnectorXls.COL_OUTER_RPORT]]
            connectors.append(connector)
            self._logger.debug("ShortName: %s" % connector.short_name)

    def readAssemblySwConnectors(self, sheet: Worksheet, swc: str, start_row: int, column_list: Dict[str, int]):
        connectors = self.getSwConnectorList(swc)
        for row in sheet.iter_rows(min_row = start_row, values_only= True):
            connector = AssemblySwConnectorData()
            connector.short_name = row[column_list[ConnectorXls.COL_SHORT_NAME]]
            connector.provider_swc = row[column_list[ConnectorXls.COL_PROVIDER_SW_C]]
            connector.pport = row[column_list[ConnectorXls.COL_PPORT]]
            connector.requester_swc = row[column_list[ConnectorXls.COL_REQUESTER_SW_C]]
            connector.rport = row[column_list[ConnectorXls.COL_RPORT]]
            connectors.append(connector)
            self._logger.debug("ShortName: %s" % connector.short_name)

    def parseDelegationSWConnectors(self, sheet: Worksheet, swc: str):
        self._logger.debug("Parse all DelegationSwConnector of %s" % swc)

        self.getColumnTitles(sheet, 1, self.column_delegation_sw_connectors)
        self.checkColumnTitles(self.column_delegation_sw_connectors, "Invalid DelegationSwConnectors Excel and column <%s> cannot be located.")
        self.readDelegationSwConnectors(sheet, swc, 2, self.column_delegation_sw_connectors)

    def parseAssemblySWConnectors(self, sheet: Worksheet, swc: str):
        self._logger.debug("Parse all AssemblySwConnector of %s" % swc)

        self.getColumnTitles(sheet, 1, self.column_assembly_sw_connectors)
        self.checkColumnTitles(self.column_assembly_sw_connectors, "Invalid AssemblySwConnectors Excel and column <%s> cannot be located.")
        self.readAssemblySwConnectors(sheet, swc, 2, self.column_assembly_sw_connectors)
        
    def read(self, excel_file: str):
        self._logger.info("Parse excel file <%s>" % excel_file) 

        wb = openpyxl.load_workbook(excel_file, data_only=True)

        for name in wb.sheetnames:
            m = re.match(r'(\w+)\s+-\s+(AC|DC)', name)
            if m:
                if m.group(2) == "DC":
                    self.parseDelegationSWConnectors(wb[name], m.group(1))
                elif m.group(2) == "AC":
                    self.parseAssemblySWConnectors(wb[name], m.group(1))
                else:
                    raise ValueError("Invalid sheet")
                
    def _addAssemblySwConnector(self, swc: CompositionSwComponentType, connector: AssemblySwConnectorData):
        sw_connector = swc.createAssemblySwConnector(connector.short_name)

        sw_connector.provider_iref = PPortInCompositionInstanceRef()
        sw_connector.provider_iref.context_component_ref = RefType()
        sw_connector.provider_iref.target_p_port_ref = RefType()
        sw_connector.provider_iref.context_component_ref.dest = "SW-COMPONENT-PROTOTYPE"
        sw_connector.provider_iref.context_component_ref.value = connector.provider_swc
        sw_connector.provider_iref.target_p_port_ref.dest = "P-PORT-PROTOTYPE"
        sw_connector.provider_iref.target_p_port_ref.value = connector.pport

        sw_connector.requester_iref = RPortInCompositionInstanceRef()
        sw_connector.requester_iref.context_component_ref = RefType()
        sw_connector.requester_iref.target_r_port_ref = RefType()
        sw_connector.requester_iref.context_component_ref.dest = "SW-COMPONENT-PROTOTYPE"
        sw_connector.requester_iref.context_component_ref.value = connector.requester_swc
        sw_connector.requester_iref.target_r_port_ref.dest = "R-PORT-PROTOTYPE"
        sw_connector.requester_iref.target_r_port_ref.value = connector.rport

    def _addDelegationSwConnector(self, swc: CompositionSwComponentType, connector: DelegationSwConnectorData):
        sw_connector = swc.createDelegationSwConnector(connector.short_name)
        if connector.inner_pport is not None and connector.outer_pport is not None:
            sw_connector.inner_port_iref = PPortInCompositionInstanceRef()
            sw_connector.inner_port_iref.context_component_ref = RefType()
            sw_connector.inner_port_iref.target_p_port_ref = RefType()
            sw_connector.outer_port_ref = RefType()
            sw_connector.inner_port_iref.context_component_ref.dest = "SW-COMPONENT-PROTOTYPE"
            sw_connector.inner_port_iref.context_component_ref.value = connector.inner_swc
            sw_connector.inner_port_iref.target_p_port_ref.dest = "P-PORT-PROTOTYPE"
            sw_connector.inner_port_iref.target_p_port_ref.value = connector.inner_pport
            sw_connector.outer_port_ref.dest = "P-PORT-PROTOTYPE"
            sw_connector.outer_port_ref.value = connector.outer_pport
        elif connector.inner_rport is not None and connector.outer_rport is not None:
            sw_connector.inner_port_iref = RPortInCompositionInstanceRef()
            sw_connector.inner_port_iref.context_component_ref = RefType()
            sw_connector.inner_port_iref.target_r_port_ref = RefType()
            sw_connector.outer_port_ref = RefType()
            sw_connector.inner_port_iref.context_component_ref.dest = "SW-COMPONENT-PROTOTYPE"
            sw_connector.inner_port_iref.context_component_ref.value = connector.inner_swc
            sw_connector.inner_port_iref.target_r_port_ref.dest = "R-PORT-PROTOTYPE"
            sw_connector.inner_port_iref.target_r_port_ref.value = connector.inner_rport
            sw_connector.outer_port_ref.dest = "R-PORT-PROTOTYPE"
            sw_connector.outer_port_ref.value = connector.outer_rport
        else:
            raise ValueError("Invalid DelegationSwConnector Configuration")

                
    def _updateCompositionSwComponent(self, swc: CompositionSwComponentType):
        # remove all the sw connector first
        swc.removeAllAssemblySwConnector()
        swc.removeAllDelegationSwConnector()

        connectors = self.getSwConnectorList(swc.short_name)

        for connector in connectors:
            #self._logger.info("Update %s" % connector.short_name)
            if isinstance(connector, AssemblySwConnectorData):
                self._addAssemblySwConnector(swc, connector)
            elif isinstance(connector, DelegationSwConnectorData):
                self._addDelegationSwConnector(swc, connector)
            else:
                raise ValueError("Invalid connector information")
                
    def _locateCompositionSwComponent(self, swc_name: str, parent: ARPackage):
        for swc in parent.getSwComponentTypes():
            if swc.short_name == swc_name:
                self._updateCompositionSwComponent(swc)
        for pkg in parent.getARPackages():
            self._locateCompositionSwComponent(swc_name, pkg)

    def update(self, document: AUTOSAR):
        for name in self.getCompositionSwComponentList():
            for pkg in document.getARPackages():
                self._locateCompositionSwComponent(name, pkg)

    