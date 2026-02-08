"""Tests for excel_parser module."""


import pytest
from openpyxl import Workbook

from armodel.parser.excel_parser import AbstractExcelParser


class ConcreteExcelParser(AbstractExcelParser):
    """Concrete implementation of AbstractExcelParser for testing."""
    pass


class TestAbstractExcelParserInit:
    """Tests for AbstractExcelParser initialization."""

    def test_init_creates_logger(self):
        """Test that AbstractExcelParser creates a logger."""
        parser = ConcreteExcelParser()
        assert parser._logger is not None


class TestGetColumnTitles:
    """Tests for getColumnTitles method."""

    def test_getColumnTitles_maps_single_column(self):
        """Test getColumnTitles with a single column."""
        parser = ConcreteExcelParser()
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Name")

        column_list = {"Name": -1}
        parser.getColumnTitles(ws, 1, column_list)

        assert column_list["Name"] == 0

    def test_getColumnTitles_maps_multiple_columns(self):
        """Test getColumnTitles with multiple columns."""
        parser = ConcreteExcelParser()
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Name")
        ws.cell(row=1, column=2, value="Age")
        ws.cell(row=1, column=3, value="City")

        column_list = {"Name": -1, "Age": -1, "City": -1}
        parser.getColumnTitles(ws, 1, column_list)

        assert column_list["Name"] == 0
        assert column_list["Age"] == 1
        assert column_list["City"] == 2

    def test_getColumnTitles_partial_mapping(self):
        """Test getColumnTitles with only some columns in list."""
        parser = ConcreteExcelParser()
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Name")
        ws.cell(row=1, column=2, value="Age")
        ws.cell(row=1, column=3, value="City")

        # Only map Name and City, not Age
        column_list = {"Name": -1, "City": -1}
        parser.getColumnTitles(ws, 1, column_list)

        assert column_list["Name"] == 0
        assert column_list["City"] == 2

    def test_getColumnTitles_custom_title_row(self):
        """Test getColumnTitles with a custom title row."""
        parser = ConcreteExcelParser()
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Header1")
        ws.cell(row=1, column=2, value="Header2")
        ws.cell(row=3, column=1, value="Name")
        ws.cell(row=3, column=2, value="Age")

        column_list = {"Name": -1, "Age": -1}
        parser.getColumnTitles(ws, 3, column_list)

        assert column_list["Name"] == 0
        assert column_list["Age"] == 1

    def test_getColumnTitles_not_found_columns_remain_negative_one(self):
        """Test that columns not found remain -1."""
        parser = ConcreteExcelParser()
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Name")
        ws.cell(row=1, column=2, value="Age")

        column_list = {"Name": -1, "Age": -1, "City": -1}
        parser.getColumnTitles(ws, 1, column_list)

        assert column_list["Name"] == 0
        assert column_list["Age"] == 1
        assert column_list["City"] == -1  # Not found

    def test_getColumnTitles_empty_column_list(self):
        """Test getColumnTitles with empty column list."""
        parser = ConcreteExcelParser()
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Name")

        column_list = {}
        parser.getColumnTitles(ws, 1, column_list)

        assert column_list == {}

    def test_getColumnTitles_with_empty_cells(self):
        """Test getColumnTitles handles empty cells."""
        parser = ConcreteExcelParser()
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Name")
        ws.cell(row=1, column=2, value=None)  # Empty cell
        ws.cell(row=1, column=3, value="City")

        column_list = {"Name": -1, "City": -1}
        parser.getColumnTitles(ws, 1, column_list)

        assert column_list["Name"] == 0
        assert column_list["City"] == 2


class TestCheckColumnTitles:
    """Tests for checkColumnTitles method."""

    def test_checkColumnTitles_all_found(self):
        """Test checkColumnTitles when all columns are found."""
        parser = ConcreteExcelParser()
        column_list = {"Name": 0, "Age": 1, "City": 2}

        # Should not raise
        parser.checkColumnTitles(column_list, "Missing column: %s")

    def test_checkColumnTitles_one_missing(self):
        """Test checkColumnTitles raises error for one missing column."""
        parser = ConcreteExcelParser()
        column_list = {"Name": 0, "Age": -1, "City": 2}

        with pytest.raises(ValueError, match="Missing column: Age"):
            parser.checkColumnTitles(column_list, "Missing column: %s")

    def test_checkColumnTitles_multiple_missing(self):
        """Test checkColumnTitles raises error for first missing column."""
        parser = ConcreteExcelParser()
        column_list = {"Name": 0, "Age": -1, "City": -1}

        # Should raise for the first missing column encountered
        with pytest.raises(ValueError, match="Missing column: Age"):
            parser.checkColumnTitles(column_list, "Missing column: %s")

    def test_checkColumnTitles_all_missing(self):
        """Test checkColumnTitles when all columns are missing."""
        parser = ConcreteExcelParser()
        column_list = {"Name": -1, "Age": -1, "City": -1}

        with pytest.raises(ValueError, match="Missing column: Name"):
            parser.checkColumnTitles(column_list, "Missing column: %s")

    def test_checkColumnTitles_custom_message(self):
        """Test checkColumnTitles with custom error message."""
        parser = ConcreteExcelParser()
        column_list = {"Name": 0, "Age": -1}

        with pytest.raises(ValueError, match="Column not found: Age"):
            parser.checkColumnTitles(column_list, "Column not found: %s")

    def test_checkColumnTitles_empty_list(self):
        """Test checkColumnTitles with empty column list."""
        parser = ConcreteExcelParser()
        column_list = {}

        # Should not raise when list is empty
        parser.checkColumnTitles(column_list, "Missing: %s")
